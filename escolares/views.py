"""
Vistas del módulo de Escolares (Servicios Escolares).
Validación final de documentos, integración de expediente, envío a CDMX.
"""
import io
import os
import zipfile
import json

from django.views.generic import TemplateView, ListView, View, CreateView, UpdateView, DetailView
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse, FileResponse
from django.shortcuts import redirect, get_object_or_404
from django.urls import reverse_lazy
from django.utils import timezone
from django.utils.text import slugify

from expediente.mixins import EscolaresRequeridoMixin
from expediente.models import (
    Expediente, Documento, ValidacionDocumento,
    EstadoExpediente, EstadoDocumento, EstadoValidacion, Modalidad, ESTADOS_INTEGRADOS
)
from expediente.notifications import notificar_alumno, registrar_cambio_estado, registrar_cambio_documento
from expediente.workflow import actualizar_estado_documento, verificar_avance_expediente

from administracion.models import Carrera


class CalendarioEscolaresView(EscolaresRequeridoMixin, TemplateView):
    template_name = 'escolares/calendario.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['carreras'] = Carrera.objects.filter(activa=True).order_by('nombre')
        return ctx


class DashboardEscolaresView(EscolaresRequeridoMixin, TemplateView):
    template_name = 'escolares/dashboard.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['expedientes_para_revision'] = Expediente.objects.filter(
            estado=EstadoExpediente.EN_REVISION_DOCUMENTOS
        ).select_related('alumno', 'modalidad').count()
        ctx['expedientes_para_integrar'] = Expediente.objects.filter(
            estado=EstadoExpediente.LISTO_INTEGRACION
        ).count()
        ctx['expedientes_enviados_cdmx'] = Expediente.objects.filter(
            estado=EstadoExpediente.TRAMITE_DGP
        ).count()
        ctx['expedientes_activos'] = Expediente.objects.exclude(
            estado__in=[EstadoExpediente.BORRADOR, EstadoExpediente.CONCLUIDO, EstadoExpediente.CANCELADO]
        ).count()
        ctx['expedientes_integrados'] = Expediente.objects.filter(
            estado__in=ESTADOS_INTEGRADOS
        ).count()
        ctx['expedientes_concluidos'] = Expediente.objects.filter(
            estado=EstadoExpediente.CONCLUIDO
        ).count()

        # Búsqueda y paginación
        qs = Expediente.objects.exclude(
            estado__in=[EstadoExpediente.BORRADOR, EstadoExpediente.CANCELADO]
        ).select_related('alumno', 'modalidad', 'alumno__carrera').order_by('-fecha_ultima_actualizacion')
        busqueda = self.request.GET.get('q', '').strip()
        carrera_id = self.request.GET.get('carrera', '')
        modalidad_id = self.request.GET.get('modalidad', '')

        if busqueda:
            qs = qs.filter(
                Q(alumno__first_name__unaccent__icontains=busqueda) |
                Q(alumno__last_name__unaccent__icontains=busqueda) |
                Q(alumno__username__unaccent__icontains=busqueda) |
                Q(alumno__numero_control__unaccent__icontains=busqueda)
            )
        if carrera_id:
            qs = qs.filter(alumno__carrera_id=carrera_id)
        if modalidad_id:
            qs = qs.filter(modalidad_id=modalidad_id)

        ctx['busqueda'] = busqueda
        ctx['carrera_id'] = carrera_id
        ctx['modalidad_id'] = modalidad_id
        ctx['carreras_filter'] = Carrera.objects.filter(activa=True)
        ctx['modalidades_filter'] = Modalidad.objects.all()
        paginator = Paginator(qs, 20)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        ctx['expedientes_recientes'] = page_obj
        ctx['page_obj'] = page_obj
        ctx['is_paginated'] = page_obj.has_other_pages()
        return ctx


class ExpedienteListaEscolaresView(EscolaresRequeridoMixin, ListView):
    model = Expediente
    template_name = 'escolares/expedientes/lista.html'
    context_object_name = 'expedientes'
    paginate_by = 20

    def get_queryset(self):
        qs = Expediente.objects.exclude(
            estado=EstadoExpediente.BORRADOR
        ).select_related(
            'alumno', 'modalidad', 'alumno__carrera'
        ).order_by('-fecha_ultima_actualizacion')
        estado = self.request.GET.get('estado')
        if estado:
            qs = qs.filter(estado=estado)
        busqueda = self.request.GET.get('q', '').strip()
        if busqueda:
            qs = qs.filter(
                Q(alumno__first_name__unaccent__icontains=busqueda) |
                Q(alumno__last_name__unaccent__icontains=busqueda) |
                Q(alumno__username__unaccent__icontains=busqueda) |
                Q(alumno__numero_control__unaccent__icontains=busqueda)
            )

        carrera_id = self.request.GET.get('carrera', '')
        modalidad_id = self.request.GET.get('modalidad', '')

        if carrera_id:
            qs = qs.filter(alumno__carrera_id=carrera_id)
        if modalidad_id:
            qs = qs.filter(modalidad_id=modalidad_id)

        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['estado_filtro'] = self.request.GET.get('estado', '')
        ctx['busqueda'] = self.request.GET.get('q', '')
        ctx['estados'] = EstadoExpediente.choices

        ctx['carrera_id'] = self.request.GET.get('carrera', '')
        ctx['modalidad_id'] = self.request.GET.get('modalidad', '')
        ctx['carreras_filter'] = Carrera.objects.filter(activa=True)
        ctx['modalidades_filter'] = Modalidad.objects.all()

        return ctx


class ExpedienteDetalleEscolaresView(EscolaresRequeridoMixin, DetailView):
    model = Expediente
    template_name = 'escolares/expedientes/detalle.html'
    context_object_name = 'expediente'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        expediente = self.object
        ctx['documentos'] = expediente.documentos.select_related(
            'tipo_documento'
        ).prefetch_related('validaciones__validado_por').order_by('tipo_documento__orden')
        ctx['historial'] = expediente.historial.select_related('realizado_por')[:15]
        return ctx



class IntegrarExpedienteView(EscolaresRequeridoMixin, View):
    """Servicios Escolares integra expediente tras aprobar todos los documentos."""

    def post(self, request, pk):
        expediente = get_object_or_404(Expediente, pk=pk)
        if expediente.estado != EstadoExpediente.EN_REVISION_DOCUMENTOS:
            messages.error(request, 'El expediente no está en revisión de documentos.')
            return redirect('escolares:expediente_detalle', pk=pk)
        # Intentar integrar (avanzar estado) y notificar al alumno
        verificar_avance_expediente(expediente)
        messages.success(request, 'Expediente integrado. Notificación enviada al alumno.')
        return redirect('escolares:expediente_detalle', pk=pk)

class ValidarDocumentoEscolaresView(EscolaresRequeridoMixin, View):
    """Servicios Escolares aprueba, rechaza o solicita corrección en un documento."""

    def post(self, request, pk):
        documento = get_object_or_404(Documento, pk=pk)
        accion = request.POST.get('accion')  # APROBAR, RECHAZAR, CORRECCION
        observaciones = request.POST.get('observaciones', '').strip()

        estado_map = {
            'APROBAR': EstadoValidacion.APROBADO,
            'RECHAZAR': EstadoValidacion.RECHAZADO,
            'CORRECCION': EstadoValidacion.REQUIERE_CORRECCION,
        }
        if accion not in estado_map:
            messages.error(request, 'Acción no válida.')
            return redirect('escolares:expediente_detalle', pk=documento.expediente.pk)

        # Enforce sequential validation (DIV -> SE)
        if not documento.puede_escolares_validar():
            messages.error(request, 'Este documento no puede ser validado por Escolares aún. Requiere visto bueno de División de Estudios.')
            return redirect('escolares:expediente_detalle', pk=documento.expediente.pk)

        validacion, _ = ValidacionDocumento.objects.get_or_create(
            documento=documento,
            departamento='ESCOLARES',
        )
        validacion.estado = estado_map[accion]
        validacion.validado_por = request.user
        validacion.observaciones = observaciones
        if not validacion.fecha_primera_revision:
            validacion.fecha_primera_revision = timezone.now()
        validacion.save()

        # Actualizar estado del documento usando lógica compartida (requiere AMBOS departamentos)
        actualizar_estado_documento(documento, realizado_por=request.user)

        if accion == 'APROBAR':
            if documento.estado == EstadoDocumento.APROBADO:
                tipo_notif = 'APROBADO'
                msg_alumno = f'El documento "{documento.tipo_documento.nombre}" ha sido aprobado por todos los departamentos.'
            else:
                tipo_notif = 'INFO'
                msg_alumno = f'Servicios Escolares aprobó el documento "{documento.tipo_documento.nombre}". Pendiente revisión de División de Estudios.'
        elif accion == 'RECHAZAR':
            tipo_notif = 'RECHAZADO'
            msg_alumno = f'El documento "{documento.tipo_documento.nombre}" fue rechazado por Servicios Escolares. Observaciones: {observaciones}'
        else:
            tipo_notif = 'CORRECCION'
            msg_alumno = f'El documento "{documento.tipo_documento.nombre}" requiere correcciones. Observaciones: {observaciones}'

        participio = {
            'APROBAR': 'aprobado',
            'RECHAZAR': 'rechazado',
            'CORRECCION': 'marcado para corrección'
        }.get(accion, accion.lower())

        registrar_cambio_documento(
            documento=documento,
            accion=f'Servicios Escolares: {accion}',
            realizado_por=request.user,
            observaciones=observaciones,
            departamento='ESCOLARES'
        )

        notificar_alumno(
            expediente=documento.expediente,
            tipo=tipo_notif,
            titulo=f'Documento {participio} por Servicios Escolares',
            mensaje=msg_alumno,
        )

        # Verificar si el expediente puede avanzar (todos los docs aprobados por ambos)
        verificar_avance_expediente(documento.expediente)

        # Si el expediente pasó a LISTO_INTEGRACION, notificar al alumno para entregar papeles originales
        if documento.expediente.estado == EstadoExpediente.LISTO_INTEGRACION:
            notificar_alumno(
                expediente=documento.expediente,
                tipo='AVANCE',
                titulo='Entrega de Papeles Originales Necesaria',
                mensaje='Todos tus documentos digitales han sido aprobados. Por favor, entrega los papeles originales en Servicios Escolares para continuar con el proceso.',
            )

        messages.success(request, f'Documento {participio}.')
        return redirect('escolares:expediente_detalle', pk=documento.expediente.pk)



class MarcarPapelesRecibidosView(EscolaresRequeridoMixin, View):
    """Escolares marca que recibió los papeles originales. Transita a PAGO_PENDIENTE."""

    def post(self, request, pk):
        expediente = get_object_or_404(Expediente, pk=pk)

        if not expediente.todos_documentos_aprobados():
            messages.error(request, 'No todos los documentos digitales están aprobados.')
            return redirect('escolares:expediente_detalle', pk=pk)

        # Verificar entrega física de fotografías
        if not expediente.foto_fisica_division or not expediente.foto_fisica_escolares:
            messages.error(request, 'Falta confirmar la entrega física de la fotografía en todos los departamentos.')
            return redirect('escolares:expediente_detalle', pk=pk)

        # Registramos primero el paso intermedio "RECIBI_PAPEL_ORIGINAL" en el historial
        registrar_cambio_estado(
            expediente=expediente,
            estado_nuevo=EstadoExpediente.RECIBI_PAPEL_ORIGINAL,
            realizado_por=request.user,
            descripcion='Servicios Escolares confirmó la integración de los papeles originales.'
        )

        expediente.estado = EstadoExpediente.PAGO_PENDIENTE
        expediente.save(update_fields=['estado', 'fecha_ultima_actualizacion'])

        registrar_cambio_estado(
            expediente=expediente,
            estado_nuevo=EstadoExpediente.PAGO_PENDIENTE,
            realizado_por=request.user,
            descripcion='El expediente avanza a etapa de pago de titulación.'
        )
        notificar_alumno(
            expediente=expediente,
            tipo='AVANCE',
            titulo='Papeles Originales Integrados — Pendiente de Pago',
            mensaje='Servicios Escolares ha integrado tus papeles originales. Por favor, sube tu comprobante de pago para continuar.',
        )
        messages.success(request, 'Papeles integrados confirmados. El expediente está en etapa de pago.')
        return redirect('escolares:expediente_detalle', pk=pk)


class ValidarPagoEscolaresView(EscolaresRequeridoMixin, View):
    """Servicios Escolares aprueba o rechaza el comprobante de pago cargado."""

    def post(self, request, pk):
        expediente = get_object_or_404(Expediente, pk=pk)
        accion = request.POST.get('accion')  # APROBAR, RECHAZAR
        observaciones = request.POST.get('observaciones', '').strip()

        if expediente.estado != EstadoExpediente.PAGO_EN_REVISION:
            messages.error(request, 'El expediente no se encuentra en revisión de pago.')
            return redirect('escolares:expediente_detalle', pk=pk)

        if accion == 'APROBAR':
            expediente.pago_validado = 'APROBADO'
            expediente.estado = EstadoExpediente.ESPERANDO_CONSTANCIA
            expediente.pago_observaciones = ''
            expediente.fecha_validacion_pago = timezone.now()
            expediente.save(update_fields=[
                'pago_validado', 'estado', 'pago_observaciones', 'fecha_validacion_pago', 'fecha_ultima_actualizacion'
            ])

            registrar_cambio_estado(
                expediente=expediente,
                estado_nuevo=EstadoExpediente.ESPERANDO_CONSTANCIA,
                realizado_por=request.user,
                descripcion='Servicios Escolares aprobó el comprobante de pago.'
            )
            notificar_alumno(
                expediente=expediente,
                tipo='APROBADO',
                titulo='Pago de Titulación Aprobado',
                mensaje='Tu comprobante de pago ha sido aprobado. Escolares subirá próximamente tu constancia de no inconveniencia.',
            )
            messages.success(request, 'Comprobante de pago aprobado con éxito. El expediente avanza a la etapa de Esperando Constancia.')

        elif accion == 'RECHAZAR':
            if not observaciones:
                messages.error(request, 'Debes escribir observaciones/motivo del rechazo.')
                return redirect('escolares:expediente_detalle', pk=pk)

            expediente.pago_validado = 'RECHAZADO'
            expediente.estado = EstadoExpediente.PAGO_PENDIENTE
            expediente.pago_observaciones = observaciones
            expediente.fecha_validacion_pago = timezone.now()
            expediente.save(update_fields=[
                'pago_validado', 'estado', 'pago_observaciones', 'fecha_validacion_pago', 'fecha_ultima_actualizacion'
            ])

            registrar_cambio_estado(
                expediente=expediente,
                estado_nuevo=EstadoExpediente.PAGO_PENDIENTE,
                realizado_por=request.user,
                descripcion=f'Servicios Escolares rechazó el comprobante de pago. Motivo: {observaciones}'
            )
            notificar_alumno(
                expediente=expediente,
                tipo='RECHAZADO',
                titulo='Comprobante de Pago Rechazado',
                mensaje=f'Tu comprobante de pago fue rechazado. Observaciones: {observaciones}',
            )
            messages.success(request, 'Comprobante de pago rechazado con éxito.')

        else:
            messages.error(request, 'Acción no válida.')

        return redirect('escolares:expediente_detalle', pk=pk)


class DescargarExpedienteView(EscolaresRequeridoMixin, View):
    """Descarga un ZIP con todos los documentos del expediente para envío físico."""

    def get(self, request, pk):
        expediente = get_object_or_404(Expediente, pk=pk)

        # Crear ZIP en memoria
        buffer = io.BytesIO()
        alumno_name = expediente.alumno.get_full_name().replace(' ', '_')
        ncontrol = expediente.alumno.username

        with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            # Agregar cada documento del expediente
            for doc in expediente.documentos.select_related('tipo_documento').all():
                if doc.archivo:
                    try:
                        archivo_path = doc.archivo.path
                        if os.path.exists(archivo_path):
                            # Nombre ASCII-safe para el archivo dentro del ZIP
                            ext = os.path.splitext(doc.archivo.name)[1].lower()
                            # Quitamos acentos y caracteres raros del nombre del documento
                            nombre_safe = slugify(doc.tipo_documento.nombre)
                            nombre_archivo = f"{doc.tipo_documento.orden:02d}_{nombre_safe}_v{doc.version}{ext}"
                            zf.write(archivo_path, nombre_archivo)
                    except Exception:
                        continue

            # Agregar fotografía digital si existe
            if expediente.fotografia_digital:
                try:
                    foto_path = expediente.fotografia_digital.path
                    if os.path.exists(foto_path):
                        ext = os.path.splitext(expediente.fotografia_digital.name)[1].lower()
                        zf.write(foto_path, f"fotografia-digital{ext}")
                except Exception:
                    pass

            # Agregar un resumen en texto
            resumen = f"""EXPEDIENTE DE TITULACIÓN - INSTITUTO TECNOLÓGICO DE APIZACO
{'=' * 60}

Alumno: {expediente.alumno.get_full_name()}
N° Control: {ncontrol}
Carrera: {getattr(expediente.alumno.carrera, 'nombre', 'N/A')}
Modalidad: {expediente.modalidad.nombre if expediente.modalidad else 'N/A'}
Plan de Estudios: {expediente.modalidad.plan_estudios.nombre if expediente.modalidad else 'N/A'}
Título del trabajo: {expediente.titulo_trabajo or 'N/A'}
Empresa: {expediente.nombre_empresa or 'N/A'}
Fecha de apertura: {expediente.fecha_apertura.strftime('%d/%m/%Y %H:%M')}
Estado: {expediente.get_estado_display()}

DOCUMENTOS INCLUIDOS
{'-' * 40}
"""
            for doc in expediente.documentos.select_related('tipo_documento').all():
                estado_txt = doc.get_estado_display()
                tiene_archivo = "Sí" if doc.archivo else "No"
                resumen += f"  - {doc.tipo_documento.nombre} (v{doc.version}) — {estado_txt} — Archivo: {tiene_archivo}\n"

            resumen += f"\n\nGenerado el: {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}\n"
            zf.writestr("_RESUMEN_EXPEDIENTE.txt", resumen)

        # Preparar respuesta "Triple Layer"
        buffer.seek(0)
        alumno_slug = slugify(expediente.alumno.get_full_name())
        filename = f"Expediente_{ncontrol}_{alumno_slug}.zip"

        # Usamos octet-stream para que el navegador NO intente previsualizar nada
        # y respete el nombre del archivo adjunto.
        response = FileResponse(
            buffer,
            as_attachment=True,
            filename=filename,
            content_type='application/octet-stream'
        )
        return response


class IniciarTramiteDGPView(EscolaresRequeridoMixin, View):
    """Escolares inicia el trámite ante la DGP (Captura, Validación y Trámite)."""

    def post(self, request, pk):
        expediente = get_object_or_404(Expediente, pk=pk)

        if expediente.estado not in [EstadoExpediente.ACTA_EXENCION, EstadoExpediente.ACTO_PROGRAMADO]:
            messages.error(request, 'El expediente debe estar en Acto Programado o tener Acta de Exención para iniciar el trámite DGP.')
            return redirect('escolares:expediente_detalle', pk=pk)

        if expediente.estado == EstadoExpediente.ACTO_PROGRAMADO and not expediente.datos_dgp_confirmados:
            messages.error(request, 'El alumno aún no ha verificado y confirmado que sus datos de título sean correctos.')
            return redirect('escolares:expediente_detalle', pk=pk)

        registrar_cambio_estado(
            expediente=expediente,
            estado_nuevo=EstadoExpediente.TRAMITE_DGP,
            realizado_por=request.user,
            descripcion='Se inició el Proceso de Captura, Validación y Trámite ante DGP (espera de 50 días).'
        )
        
        notificar_alumno(
            expediente=expediente,
            tipo='AVANCE',
            titulo='Trámite de Título en DGP Iniciado',
            mensaje='Tu proceso ha avanzado a la captura y validación DGP. Revisa las instrucciones en tu panel sobre los tiempos de espera y monitoreo.',
        )
        
        messages.success(request, 'Trámite DGP iniciado exitosamente. El alumno ha sido notificado con las instrucciones.')
        return redirect('escolares:expediente_detalle', pk=pk)


class ValidarCedulaEscolaresView(EscolaresRequeridoMixin, View):
    """Escolares revisa la cédula electrónica subida por el alumno."""

    def post(self, request, pk):
        expediente = get_object_or_404(Expediente, pk=pk)
        accion = request.POST.get('accion')
        observaciones = request.POST.get('observaciones_cedula', '').strip()

        if expediente.estado != EstadoExpediente.CEDULA_EN_REVISION:
            messages.error(request, 'La cédula no está en revisión.')
            return redirect('escolares:expediente_detalle', pk=pk)

        if accion == 'APROBAR':
            # Mantener el estado o avanzar a un estado intermedio donde esperamos asignar cita
            # En nuestro flujo, pasaremos directo a asignar cita en otra vista, pero si la aprueban
            # y no asignan cita de inmediato, podemos dejarlo en CEDULA_EN_REVISION o pasarlo a un 
            # estado intermedio. Lo ideal es pedir que apruebe y asigne cita en el mismo paso.
            # Vamos a manejar esto en la vista de AgendarCitaEntregaView. Esta vista será solo para rechazo.
            pass

        elif accion == 'RECHAZAR':
            if not observaciones:
                messages.error(request, 'Debes proporcionar observaciones si rechazas la cédula.')
                return redirect('escolares:expediente_detalle', pk=pk)

            expediente.observaciones_cedula = observaciones
            expediente.save(update_fields=['observaciones_cedula'])

            registrar_cambio_estado(
                expediente=expediente,
                estado_nuevo=EstadoExpediente.CEDULA_RECHAZADA,
                realizado_por=request.user,
                descripcion=f'Cédula electrónica rechazada. Observaciones: {observaciones}'
            )

            notificar_alumno(
                expediente=expediente,
                tipo='RECHAZADO',
                titulo='Cédula Profesional Rechazada',
                mensaje=f'Hubo un problema con la cédula que subiste. Observaciones: {observaciones}. Por favor, vuelve a cargarla.',
            )
            messages.success(request, 'Cédula rechazada exitosamente.')

        return redirect('escolares:expediente_detalle', pk=pk)


class AgendarCitaEntregaView(EscolaresRequeridoMixin, View):
    """Escolares aprueba la cédula y agenda la cita de entrega física."""

    def post(self, request, pk):
        expediente = get_object_or_404(Expediente, pk=pk)
        
        if expediente.estado != EstadoExpediente.CEDULA_EN_REVISION:
            messages.error(request, 'La cédula debe estar en revisión para agendar cita.')
            return redirect('escolares:expediente_detalle', pk=pk)

        fecha_cita = request.POST.get('fecha_cita')
        instrucciones = request.POST.get('instrucciones_cita', '').strip()

        if not fecha_cita:
            messages.error(request, 'Debes proporcionar una fecha y hora para la cita.')
            return redirect('escolares:expediente_detalle', pk=pk)

        expediente.fecha_cita_entrega = fecha_cita
        expediente.instrucciones_cita = instrucciones
        expediente.save(update_fields=['fecha_cita_entrega', 'instrucciones_cita'])

        registrar_cambio_estado(
            expediente=expediente,
            estado_nuevo=EstadoExpediente.CITA_ENTREGA,
            realizado_por=request.user,
            descripcion=f'Cédula aprobada y Cita programada para {fecha_cita}.'
        )

        notificar_alumno(
            expediente=expediente,
            tipo='AVANCE',
            titulo='¡Cita de Entrega de Título Programada!',
            mensaje='Tu cédula ha sido aprobada y se ha programado tu cita para la entrega de documentos originales y título impreso. Revisa tu panel para más detalles.',
        )

        messages.success(request, 'Cita programada exitosamente.')
        return redirect('escolares:expediente_detalle', pk=pk)

class ConcluirProcesoView(EscolaresRequeridoMixin, View):
    """Escolares concluye el trámite tras entregar los documentos en la cita."""
    
    def post(self, request, pk):
        expediente = get_object_or_404(Expediente, pk=pk)
        
        if expediente.estado != EstadoExpediente.CITA_ENTREGA:
            messages.error(request, 'El expediente debe estar en Cita de Entrega para poder concluirlo.')
            return redirect('escolares:expediente_detalle', pk=pk)
            
        registrar_cambio_estado(
            expediente=expediente,
            estado_nuevo=EstadoExpediente.CONCLUIDO,
            realizado_por=request.user,
            descripcion='Proceso de titulación concluido exitosamente. Documentos y título entregados.'
        )
        
        notificar_alumno(
            expediente=expediente,
            tipo='APROBADO',
            titulo='¡Proceso de Titulación Concluido!',
            mensaje='Has recibido tu título impreso y originales. ¡Felicidades, tu proceso de titulación ha concluido exitosamente!',
        )
        
        messages.success(request, 'El proceso del alumno ha sido concluido.')
        return redirect('escolares:expediente_detalle', pk=pk)


class MarcarFotografiaEntregadaView(EscolaresRequeridoMixin, View):
    """Servicios Escolares marca la fotografía física como entregada."""
    def post(self, request, pk):
        expediente = get_object_or_404(Expediente, pk=pk)
        
        # Permitir marcar fotografía física en cualquier estado a partir de revisión inicial
        # (Quitamos la restricción restrictiva de INTEGRADO)
        
        entregada = request.POST.get('entregada') == 'on'
        expediente.foto_fisica_escolares = entregada
        expediente.save(update_fields=['foto_fisica_escolares', 'fecha_ultima_actualizacion'])
        
        action_txt = 'ENTREGADA' if entregada else 'PENDIENTE'
        messages.success(request, f'Fotografía física marcada como {action_txt} en Servicios Escolares.')
        
        # Auditoría: Intentar registrar en el documento de fotografía, si no existe, en el expediente
        from expediente.notifications import registrar_cambio_documento, registrar_cambio_estado, notificar_alumno
        foto_doc = expediente.get_documento_fotografia
        
        if foto_doc:
            registrar_cambio_documento(
                documento=foto_doc,
                accion=f'Fotografía física: {action_txt}',
                realizado_por=request.user,
                departamento='ESCOLARES'
            )
        else:
            registrar_cambio_estado(
                expediente=expediente,
                estado_nuevo=expediente.estado,  # No cambia el estado real
                realizado_por=request.user,
                descripcion=f'Fotografía física marcada como {action_txt} por Servicios Escolares.'
            )

        # Notificar al alumno
        notificar_alumno(
            expediente=expediente,
            tipo='INFO' if entregada else 'URGENTE',
            titulo=f'Fotografía física {action_txt.lower()}',
            mensaje=f'Servicios Escolares ha marcado tu fotografía física como {action_txt.lower()}.',
        )
        
        return redirect('escolares:expediente_detalle', pk=pk)



class SubirConstanciaEscolaresView(EscolaresRequeridoMixin, View):
    """Servicios Escolares sube la Constancia de No Inconveniencia firmada en PDF."""

    def post(self, request, pk):
        expediente = get_object_or_404(Expediente, pk=pk)
        archivo_pdf = request.FILES.get('constancia_pdf')

        if not archivo_pdf:
            messages.error(request, 'Debes seleccionar un archivo PDF.')
            return redirect('escolares:expediente_detalle', pk=pk)

        if not archivo_pdf.name.lower().endswith('.pdf'):
            messages.error(request, 'El archivo debe ser un PDF.')
            return redirect('escolares:expediente_detalle', pk=pk)

        # Guardar el archivo
        expediente.constancia_no_inconveniencia = archivo_pdf
        expediente.fecha_constancia = timezone.now()
        
        if expediente.estado == EstadoExpediente.ESPERANDO_CONSTANCIA:
            expediente.estado = EstadoExpediente.CONSTANCIA_EN_REVISION
            expediente.save(update_fields=['constancia_no_inconveniencia', 'fecha_constancia', 'estado'])
            
            registrar_cambio_estado(
                expediente=expediente,
                estado_nuevo=EstadoExpediente.CONSTANCIA_EN_REVISION,
                realizado_por=request.user,
                descripcion='Constancia de No Inconveniencia subida. Pendiente de revisión por División de Estudios.'
            )
        else:
            expediente.save(update_fields=['constancia_no_inconveniencia', 'fecha_constancia'])

        # Notificar a División de Estudios
        from expediente.notifications import notificar_usuarios_division
        notificar_usuarios_division(
            expediente=expediente,
            titulo='Constancia de No Inconveniencia Generada',
            mensaje=f'Servicios Escolares ha cargado la Constancia de No Inconveniencia para el alumno {expediente.alumno.get_full_name()}. Por favor revísala y valídala para continuar con el proceso de titulación.',
            url=reverse_lazy('academico:expediente_detalle', kwargs={'pk': pk})
        )

        # Notificar al alumno
        notificar_alumno(
            expediente=expediente,
            tipo='INFO',
            titulo='Constancia de No Inconveniencia Disponible',
            mensaje='Servicios Escolares ha cargado tu Constancia de No Inconveniencia. División de Estudios la revisará pronto.',
        )

        messages.success(request, 'Constancia de No Inconveniencia subida exitosamente y enviada a revisión.')
        return redirect('escolares:expediente_detalle', pk=pk)


class EnviarNotificacionDGPView(EscolaresRequeridoMixin, View):
    """Escolares envía la notificación DGP al alumno sin cambiar de etapa."""

    def post(self, request, pk):
        expediente = get_object_or_404(Expediente, pk=pk)

        if expediente.estado != EstadoExpediente.ACTO_PROGRAMADO:
            messages.error(request, 'El expediente debe estar en Acto Programado para enviar esta notificación.')
            return redirect('escolares:expediente_detalle', pk=pk)

        expediente.notificacion_dgp_enviada = True
        expediente.save(update_fields=['notificacion_dgp_enviada'])

        mensaje_dgp = (
            'DGP_INSTRUCCIONES\n\n'
            'PROCESO DE CAPTURA, VALIDACIÓN Y TRÁMITE\n\n'
            'a) Captura, durante un tiempo transcurrido de 10 días hábiles, a partir de la fecha de protocolo, '
            'ingresar a la plataforma que indica el Estatus del Título (Validación Títulos):\n'
            'https://etitulos.tecnm.mx/validacion\n'
            'Una vez que verifiques los datos confirma que los datos son correctos enviando un mensaje a servicios escolares.\n\n'
            'b) Validar y revisar, realizando el monitoreo durante los 50 días hábiles, tiempo que dura el proceso '
            'en captura de información, revisión y expedición del documento Título Profesional, por la Dirección '
            'General de Profesiones (DGP), para observar el progreso del estatus del Título.\n'
            '* Si el estatus es 5. Registrado en la plataforma de títulos electrónicos de la DGP, '
            'esto te permite tramitar tu cédula profesional en el paso c.\n\n'
            'c) Tramitar tu Cédula Profesional electrónica en la siguiente plataforma:\n'
            'https://siurp.sep.gob.mx/mvc/cedulaElectronica\n'
            'Descargar el archivo y enviarlo a Servicios Escolares, para programar la cita de entrega '
            'de documentación original y Título Profesional.'
        )

        notificar_alumno(
            expediente=expediente,
            tipo='AVANCE',
            titulo='Instrucciones DGP — Proceso de Captura, Validación y Trámite',
            mensaje=mensaje_dgp,
        )
        
        # Auditoría interna
        registrar_cambio_estado(
            expediente=expediente,
            estado_nuevo=expediente.estado,
            realizado_por=request.user,
            descripcion='Notificación DGP enviada al alumno exitosamente.'
        )
        
        messages.success(request, 'Notificación enviada al alumno exitosamente.')
        return redirect('escolares:expediente_detalle', pk=pk)


class SubirActaExencionView(EscolaresRequeridoMixin, View):
    """Escolares sube el Acta de Exención / Examen, notifica al alumno y avanza la etapa."""

    def post(self, request, pk):
        expediente = get_object_or_404(Expediente, pk=pk)

        if expediente.estado not in [EstadoExpediente.ACTO_PROGRAMADO, EstadoExpediente.ACTA_EXENCION, EstadoExpediente.TRAMITE_DGP]:
            messages.error(request, 'El expediente debe estar en Acto Programado, Acta de Exención o Trámite DGP para subir el acta.')
            return redirect('escolares:expediente_detalle', pk=pk)

        if not expediente.datos_dgp_confirmados:
            messages.error(request, 'El alumno aún no ha verificado y confirmado que sus datos de título sean correctos.')
            return redirect('escolares:expediente_detalle', pk=pk)

        archivo_pdf = request.FILES.get('acta_pdf')
        if not archivo_pdf:
            messages.error(request, 'Debes seleccionar un archivo PDF para el acta.')
            return redirect('escolares:expediente_detalle', pk=pk)

        if not archivo_pdf.name.lower().endswith('.pdf'):
            messages.error(request, 'El archivo debe ser un PDF.')
            return redirect('escolares:expediente_detalle', pk=pk)

        expediente.acta_exencion_pdf = archivo_pdf
        estado_anterior = expediente.estado
        if expediente.estado == EstadoExpediente.ACTO_PROGRAMADO:
            expediente.estado = EstadoExpediente.ACTA_EXENCION
        expediente.save(update_fields=['acta_exencion_pdf', 'estado', 'fecha_ultima_actualizacion'])

        registrar_cambio_estado(
            expediente=expediente,
            estado_nuevo=expediente.estado,
            realizado_por=request.user,
            descripcion=f'Acta de Exención/Examen subida exitosamente. Estado anterior: {estado_anterior}.'
        )

        notificar_alumno(
            expediente=expediente,
            tipo='AVANCE',
            titulo='Acta de Exención / Examen Disponible',
            mensaje='Tu Acta de Exención de Examen Profesional (y/o Acta de Examen) ha sido generada y está disponible en el sistema. Si requieres el documento físico, favor de pasar a Servicios Escolares por él.',
        )

        messages.success(request, f'Acta subida y enviada al alumno. El expediente está en etapa: {expediente.get_estado_display()}.')
        return redirect('escolares:expediente_detalle', pk=pk)


class EnviarRecordatorioEscolaresView(EscolaresRequeridoMixin, View):
    """Envía un recordatorio (email/notificación) al alumno para que continúe su trámite."""

    def post(self, request, pk):
        expediente = get_object_or_404(Expediente, pk=pk)
        
        estados_pausa = [
            EstadoExpediente.EN_CORRECCION,
            EstadoExpediente.DOCUMENTOS_PENDIENTES,
            EstadoExpediente.PAGO_PENDIENTE,
            EstadoExpediente.CEDULA_RECHAZADA,
            EstadoExpediente.EMPASTADO_PENDIENTE,
        ]

        if expediente.estado not in estados_pausa:
            messages.warning(request, 'El expediente no se encuentra en un estado que requiera acción del alumno.')
            return redirect('escolares:expediente_detalle', pk=pk)

        # Evitar spam (solo 1 recordatorio cada 24 hrs)
        if expediente.fecha_ultimo_recordatorio:
            diff = timezone.now() - expediente.fecha_ultimo_recordatorio
            if diff.total_seconds() < 86400:  # 24 horas
                messages.warning(request, 'Ya se envió un recordatorio a este alumno en las últimas 24 horas.')
                return redirect('escolares:expediente_detalle', pk=pk)

        # Configurar mensaje según estado
        if expediente.estado == EstadoExpediente.PAGO_PENDIENTE:
            motivo = 'subir tu comprobante de pago de titulación'
        elif expediente.estado == EstadoExpediente.EMPASTADO_PENDIENTE:
            motivo = 'entregar tu trabajo empastado físico en División de Estudios'
        elif expediente.estado == EstadoExpediente.CEDULA_RECHAZADA:
            motivo = 'corregir y subir tu Cédula Profesional en formato PDF'
        else:
            motivo = 'cargar o corregir los documentos pendientes de tu expediente'

        mensaje = f'Servicios Escolares te recuerda que debes {motivo} para poder continuar con tu trámite de titulación. Por favor, atiende este requerimiento a la brevedad posible.'

        notificar_alumno(
            expediente=expediente,
            tipo='URGENTE',
            titulo='Recordatorio de Trámite Pendiente',
            mensaje=mensaje,
        )

        expediente.fecha_ultimo_recordatorio = timezone.now()
        expediente.save(update_fields=['fecha_ultimo_recordatorio'])

        messages.success(request, 'Recordatorio enviado exitosamente al alumno.')
        return redirect('escolares:expediente_detalle', pk=pk)


class ExportarExpedientesExcelView(EscolaresRequeridoMixin, View):
    """Exporta la lista de expedientes filtrados a un archivo Excel."""

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        # 1. Obtener los datos con los mismos filtros que la vista de lista
        qs = Expediente.objects.exclude(
            estado=EstadoExpediente.BORRADOR
        ).select_related(
            'alumno', 'modalidad', 'alumno__carrera'
        ).order_by('-fecha_ultima_actualizacion')
        
        estado = request.GET.get('estado')
        if estado:
            qs = qs.filter(estado=estado)
            
        busqueda = request.GET.get('q', '').strip()
        if busqueda:
            qs = qs.filter(
                Q(alumno__first_name__unaccent__icontains=busqueda) |
                Q(alumno__last_name__unaccent__icontains=busqueda) |
                Q(alumno__username__unaccent__icontains=busqueda) |
                Q(alumno__numero_control__unaccent__icontains=busqueda)
            )

        carrera_id = request.GET.get('carrera', '')
        if carrera_id:
            qs = qs.filter(alumno__carrera_id=carrera_id)
            
        modalidad_id = request.GET.get('modalidad', '')
        if modalidad_id:
            qs = qs.filter(modalidad_id=modalidad_id)

        # 2. Crear el libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Expedientes"

        # 3. Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1B396A", end_color="1B396A", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        headers = [
            "ID", "N° de Control", "Alumno", "Carrera", "Modalidad", 
            "Estado Actual", "Fecha de Apertura", "Última Actualización"
        ]

        # Escribir encabezados
        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header_title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Escribir datos
        for row_num, exp in enumerate(qs, 2):
            # Formatear fechas de manera segura
            fecha_aper = exp.fecha_apertura.strftime("%d/%m/%Y %H:%M") if exp.fecha_apertura else ""
            fecha_act = exp.fecha_ultima_actualizacion.strftime("%d/%m/%Y %H:%M") if exp.fecha_ultima_actualizacion else ""
            
            carrera_nombre = exp.alumno.carrera.nombre if hasattr(exp.alumno, 'carrera') and exp.alumno.carrera else "N/A"
            modalidad_nombre = exp.modalidad.nombre if exp.modalidad else "N/A"

            row_data = [
                exp.pk,
                exp.alumno.username,
                exp.alumno.get_full_name(),
                carrera_nombre,
                modalidad_nombre,
                exp.get_estado_display(),
                fecha_aper,
                fecha_act
            ]

            for col_num, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                cell.border = thin_border
                if col_num in [1, 2, 7, 8]:  # ID, Control y fechas centrados
                    cell.alignment = Alignment(horizontal="center")

        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 35
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20

        # 4. Preparar la respuesta HTTP
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Reporte_Expedientes_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        wb.save(response)
        return response


class EstadisticasEscolaresView(EscolaresRequeridoMixin, TemplateView):
    """Vista de estadísticas generales para Servicios Escolares."""
    template_name = 'escolares/estadisticas.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        
        # Filtros
        year = self.request.GET.get('year', str(timezone.now().year))
        
        # Querysets base
        qs_base = Expediente.objects.exclude(estado=EstadoExpediente.BORRADOR)
        qs_year = qs_base.filter(fecha_apertura__year=year)
        
        # 1. Conteo de expedientes
        abiertos_total = qs_base.count()
        abiertos_year = qs_year.count()
        
        concluidos_total = qs_base.filter(estado=EstadoExpediente.CONCLUIDO).count()
        concluidos_year = qs_year.filter(estado=EstadoExpediente.CONCLUIDO).count()
        
        inconclusos_total = abiertos_total - concluidos_total - qs_base.filter(estado=EstadoExpediente.CANCELADO).count()
        inconclusos_year = abiertos_year - concluidos_year - qs_year.filter(estado=EstadoExpediente.CANCELADO).count()

        ctx['stats'] = {
            'abiertos_total': abiertos_total,
            'abiertos_year': abiertos_year,
            'concluidos_total': concluidos_total,
            'concluidos_year': concluidos_year,
            'inconclusos_total': inconclusos_total,
            'inconclusos_year': inconclusos_year,
        }
        
        # 2. Histórico anual (últimos 5 años)
        current_year = timezone.now().year
        historico = []
        for y in range(current_year - 4, current_year + 1):
            qs_y = qs_base.filter(fecha_apertura__year=y)
            historico.append({
                'year': y,
                'abiertos': qs_y.count(),
                'concluidos': qs_y.filter(estado=EstadoExpediente.CONCLUIDO).count(),
            })
        ctx['historico_anual'] = sorted(historico, key=lambda x: x['year'], reverse=True)
        
        # 3. Listado de alumnos con proceso inconcluso (filtrable por año)
        qs_inconclusos = qs_year.exclude(
            estado__in=[EstadoExpediente.CONCLUIDO, EstadoExpediente.CANCELADO]
        ).select_related('alumno', 'alumno__carrera', 'modalidad').order_by('fecha_apertura')
        
        # Paginación de inconclusos
        paginator = Paginator(qs_inconclusos, 20)
        page_number = self.request.GET.get('page')
        ctx['page_obj'] = paginator.get_page(page_number)
        
        # Años disponibles para el filtro
        years = Expediente.objects.dates('fecha_apertura', 'year', order='DESC')
        ctx['years'] = [d.year for d in years]
        if not ctx['years']:
            ctx['years'] = [current_year]
        ctx['selected_year'] = int(year) if year.isdigit() else current_year

        # 4. Datos para gráficos interactivos (ApexCharts)
        from django.db.models import Count

        # Carreras con mayor índice de titulación (Titulados: estado CONCLUIDO)
        carreras_qs = qs_base.filter(estado=EstadoExpediente.CONCLUIDO).values('alumno__carrera__nombre').annotate(total=Count('id')).order_by('-total')[:10]
        carreras_nombres = [item['alumno__carrera__nombre'] if item['alumno__carrera__nombre'] else "N/A" for item in carreras_qs]
        carreras_totales = [item['total'] for item in carreras_qs]

        ctx['carreras_chart'] = {
            'labels': json.dumps(carreras_nombres),
            'values': json.dumps(carreras_totales),
        }
        
        # Distribución por Género (Alumnos en expedientes)
        femenino_count = qs_base.filter(alumno__genero='F').count()
        masculino_count = qs_base.filter(alumno__genero='M').count()
        total_gender = femenino_count + masculino_count
        
        femenino_pct = round((femenino_count / total_gender * 100), 1) if total_gender > 0 else 0
        masculino_pct = round((masculino_count / total_gender * 100), 1) if total_gender > 0 else 0
        
        ctx['genero_stats'] = {
            'femenino_count': femenino_count,
            'masculino_count': masculino_count,
            'femenino_pct': femenino_pct,
            'masculino_pct': masculino_pct,
            'total': total_gender,
        }
        
        return ctx


class ExportarEstadisticasDatosExcelView(EscolaresRequeridoMixin, View):
    """Exporta estadísticas o lista de alumnos inconclusos a formato Excel."""

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from django.http import HttpResponse
        
        tipo = request.GET.get('tipo', 'metricas')
        year = request.GET.get('year', str(timezone.now().year))
        
        qs_base = Expediente.objects.exclude(estado=EstadoExpediente.BORRADOR)
        qs_year = qs_base.filter(fecha_apertura__year=year)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1B396A", end_color="1B396A", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        
        if tipo == 'alumnos':
            # Exportar alumnos inconclusos
            ws.title = "Alumnos Inconclusos"
            
            headers = ["No. Control", "Nombre Completo", "Carrera", "Modalidad", "Estado Actual", "Fecha de Apertura"]
            for col_num, header_title in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header_title)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
                
            qs_inconclusos = qs_year.exclude(
                estado__in=[EstadoExpediente.CONCLUIDO, EstadoExpediente.CANCELADO]
            ).select_related('alumno', 'alumno__carrera', 'modalidad').order_by('fecha_apertura')
            
            for row_num, exp in enumerate(qs_inconclusos, 2):
                ws.cell(row=row_num, column=1, value=exp.alumno.username).border = thin_border
                ws.cell(row=row_num, column=2, value=exp.alumno.get_full_name()).border = thin_border
                ws.cell(row=row_num, column=3, value=exp.alumno.carrera.nombre if exp.alumno.carrera else "N/A").border = thin_border
                ws.cell(row=row_num, column=4, value=exp.modalidad.nombre if exp.modalidad else "N/A").border = thin_border
                ws.cell(row=row_num, column=5, value=exp.get_estado_display()).border = thin_border
                ws.cell(row=row_num, column=6, value=exp.fecha_apertura.strftime("%d/%m/%Y") if exp.fecha_apertura else "").border = thin_border
                
            # Auto-fit columns
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="Alumnos_Inconclusos_{year}.xlsx"'
            wb.save(response)
            return response
            
        else:
            # Exportar métricas generales
            ws.title = "Resumen de Estadísticas"
            
            # KPI Cards
            ws.cell(row=1, column=1, value="Indicador").font = header_font
            ws.cell(row=1, column=1).fill = header_fill
            ws.cell(row=1, column=2, value="Año Seleccionado").font = header_font
            ws.cell(row=1, column=2).fill = header_fill
            ws.cell(row=1, column=3, value="Histórico Total").font = header_font
            ws.cell(row=1, column=3).fill = header_fill
            
            abiertos_total = qs_base.count()
            abiertos_year = qs_year.count()
            concluidos_total = qs_base.filter(estado=EstadoExpediente.CONCLUIDO).count()
            concluidos_year = qs_year.filter(estado=EstadoExpediente.CONCLUIDO).count()
            inconclusos_total = abiertos_total - concluidos_total - qs_base.filter(estado=EstadoExpediente.CANCELADO).count()
            inconclusos_year = abiertos_year - concluidos_year - qs_year.filter(estado=EstadoExpediente.CANCELADO).count()
            
            metrics = [
                ("Expedientes Abiertos", abiertos_year, abiertos_total),
                ("Procesos Concluidos", concluidos_year, concluidos_total),
                ("Procesos Inconclusos", inconclusos_year, inconclusos_total)
            ]
            
            for row_num, (name, val_y, val_t) in enumerate(metrics, 2):
                ws.cell(row=row_num, column=1, value=name).border = thin_border
                ws.cell(row=row_num, column=2, value=val_y).border = thin_border
                ws.cell(row=row_num, column=3, value=val_t).border = thin_border
                
            # Gender distribution
            femenino_count = qs_base.filter(alumno__genero='F').count()
            masculino_count = qs_base.filter(alumno__genero='M').count()
            total_gender = femenino_count + masculino_count
            femenino_pct = round((femenino_count / total_gender * 100), 1) if total_gender > 0 else 0
            masculino_pct = round((masculino_count / total_gender * 100), 1) if total_gender > 0 else 0
            
            ws.cell(row=6, column=1, value="Distribución por Género").font = Font(bold=True)
            ws.cell(row=7, column=1, value="Género").font = header_font
            ws.cell(row=7, column=1).fill = header_fill
            ws.cell(row=7, column=2, value="Cantidad").font = header_font
            ws.cell(row=7, column=2).fill = header_fill
            ws.cell(row=7, column=3, value="Porcentaje").font = header_font
            ws.cell(row=7, column=3).fill = header_fill
            
            ws.cell(row=8, column=1, value="Femenino").border = thin_border
            ws.cell(row=8, column=2, value=femenino_count).border = thin_border
            ws.cell(row=8, column=3, value=f"{femenino_pct}%").border = thin_border
            
            ws.cell(row=9, column=1, value="Masculino").border = thin_border
            ws.cell(row=9, column=2, value=masculino_count).border = thin_border
            ws.cell(row=9, column=3, value=f"{masculino_pct}%").border = thin_border
            
            ws.cell(row=10, column=1, value="Total").border = thin_border
            ws.cell(row=10, column=2, value=total_gender).border = thin_border
            ws.cell(row=10, column=3, value="100%").border = thin_border
            
            # History
            ws.cell(row=12, column=1, value="Histórico Anual (Últimos 5 años)").font = Font(bold=True)
            ws.cell(row=13, column=1, value="Año").font = header_font
            ws.cell(row=13, column=1).fill = header_fill
            ws.cell(row=13, column=2, value="Abiertos").font = header_font
            ws.cell(row=13, column=2).fill = header_fill
            ws.cell(row=13, column=3, value="Concluidos").font = header_font
            ws.cell(row=13, column=3).fill = header_fill
            
            current_year = timezone.now().year
            row_idx = 14
            for y in range(current_year - 4, current_year + 1):
                qs_y = qs_base.filter(fecha_apertura__year=y)
                ws.cell(row=row_idx, column=1, value=y).border = thin_border
                ws.cell(row=row_idx, column=2, value=qs_y.count()).border = thin_border
                ws.cell(row=row_idx, column=3, value=qs_y.filter(estado=EstadoExpediente.CONCLUIDO).count()).border = thin_border
                row_idx += 1
                
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 16)
                
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="Estadisticas_Titulacion_{year}.xlsx"'
            wb.save(response)
            return response
