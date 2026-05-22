"""Script para exportar Carreras y Departamentos a Excel."""
import os
import sys
import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'titulacion.settings')
sys.path.insert(0, os.path.dirname(__file__))
django.setup()

from administracion.models import Carrera, Departamento

try:
    import openpyxl
except ImportError:
    print("Instalando openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
    import openpyxl

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()

# --- Hoja de Departamentos ---
ws_dep = wb.active
ws_dep.title = "Departamentos"

header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=12)
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

dep_headers = ['ID', 'Clave', 'Nombre del Departamento', 'Rol Responsable']
for col, h in enumerate(dep_headers, 1):
    cell = ws_dep.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

for row_idx, dep in enumerate(Departamento.objects.all().order_by('id'), 2):
    data = [dep.id, dep.clave, dep.nombre, dep.get_rol_responsable_display()]
    for col, val in enumerate(data, 1):
        cell = ws_dep.cell(row=row_idx, column=col, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center')

ws_dep.column_dimensions['A'].width = 6
ws_dep.column_dimensions['B'].width = 15
ws_dep.column_dimensions['C'].width = 50
ws_dep.column_dimensions['D'].width = 35

# --- Hoja de Carreras ---
ws_car = wb.create_sheet("Carreras")

car_headers = ['ID', 'Clave', 'Nombre de la Carrera', 'Activa', 'Departamento']
for col, h in enumerate(car_headers, 1):
    cell = ws_car.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

for row_idx, car in enumerate(Carrera.objects.select_related('departamento').all().order_by('id'), 2):
    data = [
        car.id,
        car.clave,
        car.nombre,
        'Si' if car.activa else 'No',
        car.departamento.nombre if car.departamento else 'Sin departamento'
    ]
    for col, val in enumerate(data, 1):
        cell = ws_car.cell(row=row_idx, column=col, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center')

ws_car.column_dimensions['A'].width = 6
ws_car.column_dimensions['B'].width = 15
ws_car.column_dimensions['C'].width = 55
ws_car.column_dimensions['D'].width = 10
ws_car.column_dimensions['E'].width = 50

output_path = os.path.join(os.path.dirname(__file__), 'carreras_departamentos.xlsx')
wb.save(output_path)
print(f"Excel guardado en: {output_path}")
