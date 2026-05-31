"""
Vistas de Importación y Exportación Masiva vía Excel para el Sistema de Titulacion ITA.
Proporciona lógica de generación de plantillas individuales/grupales (vacías o con datos)
y un procesador transaccional atómico tipo Upsert para evitar la duplicación de registros.
"""
import csv
import io
import zipfile
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from django.views.generic import TemplateView, View
from django.db import transaction
from django.shortcuts import render, redirect
from django.urls import reverse
from django.contrib import messages
from django.http import HttpResponse, Http404
from django.contrib.auth import get_user_model

from expediente.mixins import AdminRequeridoMixin
from administracion.models import Carrera, Departamento, Profesor, Rol, Genero, Usuario
from expediente.models import PlanEstudios, Modalidad
from alumnos.models import PerfilAlumno

Usuario = get_user_model()


class ImportarExportarHubView(AdminRequeridoMixin, TemplateView):
    """
    Vista principal de la herramienta de importación y exportación.
    Muestra pestañas separadas por catálogo y sus respectivos controles de descarga/carga.
    """
    template_name = 'administracion/importar_exportar.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        # Pasar conteos actuales de registros en el sistema para contexto informativo
        ctx['counts'] = {
            'departamentos': Departamento.objects.count(),
            'carreras': Carrera.objects.count(),
            'planes': PlanEstudios.objects.count(),
            'modalidades': Modalidad.objects.count(),
            'profesores': Profesor.objects.count(),
            'alumnos': Usuario.objects.filter(rol=Rol.ALUMNO).count(),
        }
        ctx['active_tab'] = self.request.GET.get('tab', 'todo')
        return ctx


class DescargarPlantillaView(AdminRequeridoMixin, View):
    """
    Genera y descarga plantillas Excel estructuradas de openpyxl.
    Soporta la descarga de un catálogo específico o del libro completo,
    ya sea vacío (con instrucciones) o pre-relleno con datos actuales (Upsert).
    """

    def get_data_for_key(self, key):
        rows_data = []
        if key == 'departamentos':
            for d in Departamento.objects.all().order_by('clave'):
                rows_data.append([d.clave, d.nombre, d.rol_responsable])
        elif key == 'carreras':
            for c in Carrera.objects.select_related('departamento').all().order_by('clave'):
                rows_data.append([c.clave, c.nombre, 'Si' if c.activa else 'No', c.departamento.clave if c.departamento else ''])
        elif key == 'planes':
            for p in PlanEstudios.objects.all().order_by('-nombre'):
                rows_data.append([p.nombre, p.descripcion, 'Si' if p.activo else 'No'])
        elif key == 'modalidades':
            for m in Modalidad.objects.select_related('plan_estudios').all().order_by('plan_estudios__nombre', 'clave'):
                rows_data.append([m.plan_estudios.nombre, m.clave, m.nombre, m.descripcion, 'Si' if m.activa else 'No'])
        elif key == 'profesores':
            for p in Profesor.objects.select_related('departamento').all().order_by('cedula'):
                rows_data.append([p.cedula, p.first_name, p.last_name, p.apellido_materno, p.titulo_academico, p.email, p.departamento.clave if p.departamento else '', 'Si' if p.activo else 'No'])
        elif key == 'alumnos':
            for u in Usuario.objects.filter(rol=Rol.ALUMNO).select_related('carrera').order_by('username'):
                p = getattr(u, 'perfil_alumno', None)
                rows_data.append([
                    u.username, u.first_name, u.last_name, u.apellido_materno, u.email,
                    u.carrera.clave if u.carrera else '',
                    p.plan_estudios.nombre if (p and p.plan_estudios) else '',
                    p.semestre_egreso if p else '',
                    str(p.promedio) if (p and p.promedio) else '',
                    u.telefono, u.genero, u.generacion or ''
                ])
        return rows_data

    def get(self, request):
        tipo = request.GET.get('tipo', 'todo')
        con_datos = request.GET.get('con_datos', 'false') == 'true'
        formato = request.GET.get('formato', 'excel')

        wb = openpyxl.Workbook()
        # Eliminar hoja activa por defecto
        default_sheet = wb.active
        wb.remove(default_sheet)

        # Configurar estilos corporativos (Matching ITA Azul #1B396A)
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1B396A', end_color='1B396A', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='BFBFBF'),
            right=Side(style='thin', color='BFBFBF'),
            top=Side(style='thin', color='BFBFBF'),
            bottom=Side(style='thin', color='BFBFBF')
        )
        data_font = Font(name='Calibri', size=11)
        data_align = Alignment(vertical='center')

        # Definir especificaciones de hojas y cabeceras
        sheets_def = {
            'departamentos': {
                'title': 'Departamentos',
                'headers': ['Clave', 'Nombre del Departamento', 'Rol Responsable'],
                'widths': [15, 45, 25],
                'help_rows': [
                    ['CB', 'Ciencias Básicas', 'ACADEMICO'],
                    ['CH', 'Ciencias de la Tierra', 'JEFE_PROYECTO']
                ]
            },
            'carreras': {
                'title': 'Carreras',
                'headers': ['Clave', 'Nombre de la Carrera', 'Activa (Si/No)', 'Clave Departamento'],
                'widths': [15, 50, 18, 25],
                'help_rows': [
                    ['ISC', 'Ingeniería en Sistemas Computacionales', 'Si', 'CB'],
                    ['IGE', 'Ingeniería en Gestión Empresarial', 'Si', 'CH']
                ]
            },
            'planes': {
                'title': 'Planes de Estudio',
                'headers': ['Nombre (Clave)', 'Descripción', 'Activo (Si/No)'],
                'widths': [25, 45, 18],
                'help_rows': [
                    ['ISIC-2010-224', 'Plan de estudios 2010 para ISC', 'Si'],
                    ['IIND-2010-227', 'Plan de estudios 2010 para Industrial', 'Si']
                ]
            },
            'modalidades': {
                'title': 'Modalidades',
                'headers': ['Plan de Estudios Nombre', 'Clave de Modalidad', 'Nombre de Modalidad', 'Descripción', 'Activa (Si/No)'],
                'widths': [25, 25, 35, 50, 18],
                'help_rows': [
                    ['ISIC-2010-224', 'TESIS', 'Tesis Profesional', 'Examen con defensa escrita y oral', 'Si'],
                    ['ISIC-2010-224', 'RESIDENCIA', 'Informe de Residencia Profesional', 'Reporte escrito de proyecto empresarial', 'Si']
                ]
            },
            'profesores': {
                'title': 'Profesores',
                'headers': ['Cédula Profesional', 'Nombre(s)', 'Apellido Paterno', 'Apellido Materno', 'Título Académico', 'Email', 'Clave Departamento', 'Activo (Si/No)'],
                'widths': [22, 25, 25, 25, 30, 35, 22, 18],
                'help_rows': [
                    ['14098703', 'JUAN', 'RAMOS', 'RAMOS', 'Dr. en Sistemas Computacionales', 'jramos@apizaco.tecnm.mx', 'CB', 'Si'],
                    ['12345678', 'MARIA', 'LOPEZ', 'GARCIA', 'M.C. en Administración', 'mlopez@apizaco.tecnm.mx', 'CH', 'Si']
                ]
            },
            'alumnos': {
                'title': 'Alumnos',
                'headers': ['Número de Control', 'Nombre(s)', 'Apellido Paterno', 'Apellido Materno', 'Email', 'Clave Carrera', 'Plan de Estudios Nombre', 'Semestre de Egreso', 'Promedio General', 'Teléfono', 'Género (M/F/O)', 'Generación (Año)'],
                'widths': [20, 25, 25, 25, 35, 18, 25, 25, 18, 18, 18, 18],
                'help_rows': [
                    ['20141720', 'DANIELA', 'SUAREZ', 'LOPEZ', 'L20141720@apizaco.tecnm.mx', 'ISC', 'ISIC-2010-224', 'Ago-Dic 2024', '91.50', '2411122334', 'F', '2020'],
                    ['20141721', 'CARLOS', 'PEREZ', 'GOMEZ', 'carlos.perez@gmail.com', 'ISC', 'ISIC-2010-224', 'Ene-Jun 2025', '85.40', '2415556677', 'M', '2021']
                ]
            }
        }

        # Determinar qué hojas agregar al archivo final
        active_keys = sheets_def.keys() if tipo == 'todo' else [tipo]
        if not all(k in sheets_def for k in active_keys):
            raise Http404("Catálogo no soportado.")

        if formato == 'csv':
            if tipo == 'todo':
                # Return a ZIP with multiple CSVs
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                    for key in active_keys:
                        sheet_meta = sheets_def[key]
                        csv_buffer = io.StringIO()
                        writer = csv.writer(csv_buffer)
                        
                        writer.writerow(sheet_meta['headers'])
                        if con_datos:
                            for row in self.get_data_for_key(key):
                                writer.writerow(row)
                        else:
                            for row in sheet_meta['help_rows']:
                                writer.writerow(row)
                        
                        zip_file.writestr(f"{sheet_meta['title']}.csv", csv_buffer.getvalue().encode('utf-8-sig'))
                
                response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
                filename = f'plantilla_{tipo}.zip' if not con_datos else f'datos_respaldo_{tipo}.zip'
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response
            else:
                # Return a single CSV
                response = HttpResponse(content_type='text/csv')
                response.charset = 'utf-8-sig' # Set BOM for Excel compatibility
                filename = f'plantilla_{tipo}.csv' if not con_datos else f'datos_respaldo_{tipo}.csv'
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                
                writer = csv.writer(response)
                # Ensure BOM is written
                response.write('\ufeff')
                
                key = tipo
                sheet_meta = sheets_def[key]
                writer.writerow(sheet_meta['headers'])
                
                if con_datos:
                    for row in self.get_data_for_key(key):
                        writer.writerow(row)
                else:
                    for row in sheet_meta['help_rows']:
                        writer.writerow(row)
                        
                return response

        # Logic for 'excel' format
        for key in active_keys:
            sheet_meta = sheets_def[key]
            ws = wb.create_sheet(title=sheet_meta['title'])
            ws.views.sheetView[0].showGridLines = True

            # 1. Agregar cabecera
            headers = sheet_meta['headers']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            ws.row_dimensions[1].height = 28

            # Configurar dimensiones de columna
            for col_idx, width in enumerate(sheet_meta['widths'], 1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = width

            # 2. Agregar contenido (Datos)
            if con_datos:
                rows_data = self.get_data_for_key(key)
                # Escribir filas en la hoja
                for row_idx, row_values in enumerate(rows_data, 2):
                    for col_idx, value in enumerate(row_values, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.font = data_font
                        cell.alignment = data_align
                        cell.border = thin_border
            else:
                for row_idx, row_values in enumerate(sheet_meta['help_rows'], 2):
                    for col_idx, value in enumerate(row_values, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)

        # Preparar respuesta HTTP de descarga de Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f'plantilla_{tipo}.xlsx' if not con_datos else f'datos_respaldo_{tipo}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

class SubirArchivoMasivoView(AdminRequeridoMixin, View):
    """
    Recibe el archivo Excel o CSV cargado por el administrador, valida la integridad,
    relaciones de llaves foráneas y ejecuta un Upsert atómico por transacción.
    """

    def post(self, request):
        tipo = request.POST.get('tipo', 'todo')
        uploaded_file = request.FILES.get('archivo_excel')

        if not uploaded_file:
            messages.error(request, "Por favor, selecciona un archivo para subir.")
            return redirect(reverse('administracion:importar_exportar') + f'?tab={tipo}')
            
        filename = uploaded_file.name.lower()
        is_csv = filename.endswith('.csv')
        is_excel = filename.endswith('.xlsx')

        if not (is_csv or is_excel):
            messages.error(request, "Formato de archivo inválido. Sube un archivo .xlsx o .csv")
            return redirect(reverse('administracion:importar_exportar') + f'?tab={tipo}')

        if tipo == 'todo' and is_csv:
            messages.error(request, "Para cargar todo el sistema a la vez, por favor sube un archivo Excel (.xlsx) con todas las pestañas. No se soporta CSV masivo unificado.")
            return redirect(reverse('administracion:importar_exportar') + f'?tab={tipo}')

        # Definir mapeo de pestañas
        sheet_mapping = {
            'departamentos': 'Departamentos',
            'carreras': 'Carreras',
            'planes': 'Planes de Estudio',
            'modalidades': 'Modalidades',
            'profesores': 'Profesores',
            'alumnos': 'Alumnos',
        }

        active_keys = sheet_mapping.keys() if tipo == 'todo' else [tipo]
        errors = []
        stats = {k: {'creados': 0, 'actualizados': 0} for k in sheet_mapping.keys()}
        self.newly_created_users = []

        try:
            with transaction.atomic():
                if is_csv:
                    # Logic for single CSV file
                    decoded_file = uploaded_file.read().decode('utf-8-sig').splitlines()
                    reader = csv.reader(decoded_file)
                    # Skip header
                    rows = list(reader)[1:]
                    
                    # Convert empty strings to None to match openpyxl behavior
                    normalized_rows = []
                    for row in rows:
                        normalized_rows.append([val if val != "" else None for val in row])
                        
                    self.procesar_hoja(tipo, normalized_rows, errors, stats, sheet_mapping[tipo])
                else:
                    # Logic for Excel file
                    try:
                        wb = openpyxl.load_workbook(uploaded_file, data_only=True)
                    except Exception as e:
                        raise ValidationError(f"No se pudo leer el archivo Excel. Detalles: {e}")

                    for key in active_keys:
                        sheet_name = sheet_mapping[key]

                        if sheet_name in wb.sheetnames:
                            ws = wb[sheet_name]
                        elif len(wb.sheetnames) == 1 and tipo != 'todo':
                            ws = wb.active
                        else:
                            if tipo == 'todo':
                                errors.append({
                                    'hoja': 'General', 'fila': 'N/A', 'columna': 'N/A',
                                    'error': f"Falta la pestaña '{sheet_name}' en el archivo."
                                })
                                continue
                            else:
                                raise ValueError("No se encontró la pestaña correspondiente en el archivo Excel.")

                        rows = list(ws.iter_rows(min_row=2, values_only=True))
                        self.procesar_hoja(key, rows, errors, stats, ws.title)

                if errors:
                    raise ValidationError("Errores de consistencia en datos.")

        except (ValidationError, Exception) as e:
            # Capturar fallos del rollback y renderizar con listado de errores
            ctx = {
                'counts': {
                    'departamentos': Departamento.objects.count(),
                    'carreras': Carrera.objects.count(),
                    'planes': PlanEstudios.objects.count(),
                    'modalidades': Modalidad.objects.count(),
                    'profesores': Profesor.objects.count(),
                    'alumnos': Usuario.objects.filter(rol=Rol.ALUMNO).count(),
                },
                'errors': errors,
                'active_tab': tipo,
            }
            if not errors:
                # Error inesperado de código/BD no capturado en la lista estructurada
                ctx['errors'] = [{
                    'hoja': 'Sistema / Base de Datos', 'fila': 'N/A', 'columna': 'N/A',
                    'error': f"Error crítico al guardar en la base de datos: {e}"
                }]
            return render(request, 'administracion/importar_exportar.html', ctx)

        # Enviar correos post-commit para evitar enviar correos si la transacción se revierte
        if hasattr(self, 'newly_created_users') and self.newly_created_users:
            from django.core.mail import EmailMultiAlternatives
            from django.conf import settings

            for u_data in self.newly_created_users:
                email = u_data['email']
                if not email:
                    continue
                
                subject = "[ITA Titulación] Tu cuenta ha sido creada — Datos de Acceso"
                
                context_data = {
                    'full_name': u_data['full_name'],
                    'numero_control': u_data['username'],
                    'password_clear': u_data['password']
                }
                from django.template.loader import render_to_string
                html_content = render_to_string('emails/nueva_cuenta.html', context_data)

                text_content = f"""Estimado(a) {u_data['full_name']},

Te informamos que tu cuenta de acceso para la plataforma de titulación del Instituto Tecnológico de Apizaco ha sido creada.

Datos de Acceso:
- Número de control / empleado: {u_data['username']}
- Contraseña temporal: {u_data['password']}

Puedes ingresar a la plataforma abriendo tu navegador web e introduciendo la dirección habitual de la institución.

IMPORTANTE: Por motivos de seguridad, el sistema te pedirá cambiar tu contraseña en tu primer inicio de sesión.

Instituto Tecnológico de Apizaco — TecNM.
"""
                try:
                    msg = EmailMultiAlternatives(
                        subject=subject,
                        body=text_content,
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        to=[email]
                    )
                    msg.attach_alternative(html_content, "text/html")
                    msg.send(fail_silently=True)
                except Exception:
                    pass
        

        # Armar mensaje de éxito con estadísticas descriptivas
        success_msg = "¡Importación y actualización masiva realizada con éxito!<br><ul class='mb-0'>"
        any_change = False
        for key in active_keys:
            meta = stats[key]
            sheet_title = sheet_mapping[key]
            if meta['creados'] > 0 or meta['actualizados'] > 0:
                any_change = True
                success_msg += f"<li><strong>{sheet_title}:</strong> {meta['creados']} creados, {meta['actualizados']} actualizados.</li>"

        if not any_change:
            success_msg += "<li>No se detectaron nuevos datos ni modificaciones que guardar (el archivo era idéntico).</li>"
        success_msg += "</ul>"

        messages.success(request, success_msg)
        return redirect(reverse('administracion:importar_exportar') + f'?tab={tipo}')

    def procesar_hoja(self, key, rows, errors, stats, sheet_name):
        """
        Deriva el procesamiento del renglón al método correspondiente de acuerdo a la hoja procesada.
        """
        for idx, row in enumerate(rows, 2):
            # Saltar filas completamente vacías
            if not any(val is not None for val in row):
                continue

            try:
                if key == 'departamentos':
                    self.procesar_departamento(row, idx, stats['departamentos'])
                elif key == 'carreras':
                    self.procesar_carrera(row, idx, stats['carreras'])
                elif key == 'planes':
                    self.procesar_plan(row, idx, stats['planes'])
                elif key == 'modalidades':
                    self.procesar_modalidad(row, idx, stats['modalidades'])
                elif key == 'profesores':
                    self.procesar_profesor(row, idx, stats['profesores'])
                elif key == 'alumnos':
                    self.procesar_alumno(row, idx, stats['alumnos'])
            except ValueError as e:
                errors.append({
                    'hoja': sheet_name,
                    'fila': idx,
                    'columna': 'Varios',
                    'error': str(e)
                })

    def procesar_departamento(self, row, fila, stat):
        if len(row) < 2 or not row[0] or not row[1]:
            raise ValueError("Clave y Nombre de Departamento son obligatorios.")

        clave = str(row[0]).strip().upper()
        nombre = str(row[1]).strip()
        rol_resp = str(row[2]).strip().upper() if len(row) > 2 and row[2] else 'ACADEMICO'

        # Validar rol
        roles_validos = [r[0] for r in Rol.choices]
        if rol_resp not in roles_validos:
            raise ValueError(f"Rol responsable '{rol_resp}' inválido. Opciones: {', '.join(roles_validos)}")

        dept, created = Departamento.objects.get_or_create(
            clave=clave,
            defaults={'nombre': nombre, 'rol_responsable': rol_resp}
        )
        if not created:
            # Actualización para evitar duplicados (Upsert)
            if dept.nombre != nombre or dept.rol_responsable != rol_resp:
                dept.nombre = nombre
                dept.rol_responsable = rol_resp
                dept.save()
                stat['actualizados'] += 1
        else:
            stat['creados'] += 1

    def procesar_carrera(self, row, fila, stat):
        if len(row) < 2 or not row[0] or not row[1]:
            raise ValueError("Clave y Nombre de Carrera son obligatorios.")

        clave = str(row[0]).strip().upper()
        nombre = str(row[1]).strip()
        activa = str(row[2]).strip().lower() in ['si', 'sí', 'yes', 'true', '1'] if len(row) > 2 and row[2] is not None else True
        clave_dept = str(row[3]).strip().upper() if len(row) > 3 and row[3] else None

        # Obtener departamento
        dept = None
        if clave_dept:
            try:
                dept = Departamento.objects.get(clave=clave_dept)
            except Departamento.DoesNotExist:
                raise ValueError(f"El Departamento con clave '{clave_dept}' no está registrado en el sistema.")

        carrera, created = Carrera.objects.get_or_create(
            clave=clave,
            defaults={'nombre': nombre, 'activa': activa, 'departamento': dept}
        )
        if not created:
            if carrera.nombre != nombre or carrera.activa != activa or carrera.departamento != dept:
                carrera.nombre = nombre
                carrera.activa = activa
                carrera.departamento = dept
                carrera.save()
                stat['actualizados'] += 1
        else:
            stat['creados'] += 1

    def procesar_plan(self, row, fila, stat):
        if not row[0]:
            raise ValueError("El Nombre (Clave) del Plan de Estudios es obligatorio.")

        nombre = str(row[0]).strip().upper()
        descripcion = str(row[1]).strip() if len(row) > 1 and row[1] else ''
        activo = str(row[2]).strip().lower() in ['si', 'sí', 'yes', 'true', '1'] if len(row) > 2 and row[2] is not None else True

        plan, created = PlanEstudios.objects.get_or_create(
            nombre=nombre,
            defaults={'descripcion': descripcion, 'activo': activo}
        )
        if not created:
            if plan.descripcion != descripcion or plan.activo != activo:
                plan.descripcion = descripcion
                plan.activo = activo
                plan.save()
                stat['actualizados'] += 1
        else:
            stat['creados'] += 1

    def procesar_modalidad(self, row, fila, stat):
        if len(row) < 3 or not row[0] or not row[1] or not row[2]:
            raise ValueError("Plan Estudios, Clave y Nombre de la Modalidad son obligatorios.")

        plan_nombre = str(row[0]).strip().upper()
        clave = str(row[1]).strip().upper()
        nombre = str(row[2]).strip()
        descripcion = str(row[3]).strip() if len(row) > 3 and row[3] else ''
        activa = str(row[4]).strip().lower() in ['si', 'sí', 'yes', 'true', '1'] if len(row) > 4 and row[4] is not None else True

        # Buscar plan de estudios
        try:
            plan = PlanEstudios.objects.get(nombre=plan_nombre)
        except PlanEstudios.DoesNotExist:
            raise ValueError(f"El Plan de Estudios '{plan_nombre}' no existe en el sistema.")

        modalidad, created = Modalidad.objects.get_or_create(
            clave=clave,
            plan_estudios=plan,
            defaults={'nombre': nombre, 'descripcion': descripcion, 'activa': activa}
        )
        if not created:
            if modalidad.nombre != nombre or modalidad.descripcion != descripcion or modalidad.activa != activa:
                modalidad.nombre = nombre
                modalidad.descripcion = descripcion
                modalidad.activa = activa
                modalidad.save()
                stat['actualizados'] += 1
        else:
            stat['creados'] += 1

    def procesar_profesor(self, row, fila, stat):
        if len(row) < 5 or not row[0] or not row[1] or not row[2] or not row[4]:
            raise ValueError("Cédula, Nombre, Apellido Paterno y Título Académico son obligatorios.")

        cedula = str(row[0]).strip().upper()
        first_name = str(row[1]).strip()
        last_name = str(row[2]).strip()
        apellido_materno = str(row[3]).strip() if len(row) > 3 and row[3] else ''
        titulo = str(row[4]).strip()
        email = str(row[5]).strip() if len(row) > 5 and row[5] else ''
        clave_dept = str(row[6]).strip().upper() if len(row) > 6 and row[6] else None
        activo = str(row[7]).strip().lower() in ['si', 'sí', 'yes', 'true', '1'] if len(row) > 7 and row[7] is not None else True

        # Buscar departamento
        dept = None
        if clave_dept:
            try:
                dept = Departamento.objects.get(clave=clave_dept)
            except Departamento.DoesNotExist:
                raise ValueError(f"El Departamento con clave '{clave_dept}' no existe en el sistema.")

        profesor, created = Profesor.objects.get_or_create(
            cedula=cedula,
            defaults={
                'first_name': first_name,
                'last_name': last_name,
                'apellido_materno': apellido_materno,
                'titulo_academico': titulo,
                'email': email,
                'departamento': dept,
                'activo': activo
            }
        )
        if not created:
            if (profesor.first_name != first_name or profesor.last_name != last_name or
                profesor.apellido_materno != apellido_materno or profesor.titulo_academico != titulo or
                profesor.email != email or profesor.departamento != dept or profesor.activo != activo):
                
                profesor.first_name = first_name
                profesor.last_name = last_name
                profesor.apellido_materno = apellido_materno
                profesor.titulo_academico = titulo
                profesor.email = email
                profesor.departamento = dept
                profesor.activo = activo
                profesor.save()
                stat['actualizados'] += 1
        else:
            stat['creados'] += 1

    def procesar_alumno(self, row, fila, stat):
        if len(row) < 7 or not row[0] or not row[1] or not row[2] or not row[4] or not row[5] or not row[6]:
            raise ValueError("N° Control, Nombre, Apellido Paterno, Email, Clave Carrera y Plan de Estudios son obligatorios.")

        control = str(row[0]).strip().upper()
        first_name = str(row[1]).strip()
        last_name = str(row[2]).strip()
        apellido_materno = str(row[3]).strip() if len(row) > 3 and row[3] else ''
        email = str(row[4]).strip()
        clave_carrera = str(row[5]).strip().upper()
        plan_nombre = str(row[6]).strip().upper()
        
        semestre_egreso = str(row[7]).strip() if len(row) > 7 and row[7] else ''
        promedio = row[8] if len(row) > 8 and row[8] is not None else None
        telefono = str(row[9]).strip() if len(row) > 9 and row[9] else ''
        
        genero_raw = str(row[10]).strip().upper() if len(row) > 10 and row[10] else 'O'
        genero = 'O'
        if genero_raw in ['M', 'MASCULINO', 'H', 'HOMBRE']:
            genero = Genero.MASCULINO
        elif genero_raw in ['F', 'FEMENINO', 'M', 'MUJER']:
            genero = Genero.FEMENINO

        generacion = int(row[11]) if len(row) > 11 and row[11] is not None else None

        # 1. Validar Carrera y Plan
        try:
            carrera = Carrera.objects.get(clave=clave_carrera)
        except Carrera.DoesNotExist:
            raise ValueError(f"La Carrera con clave '{clave_carrera}' no existe en el sistema.")

        try:
            plan = PlanEstudios.objects.get(nombre=plan_nombre)
        except PlanEstudios.DoesNotExist:
            raise ValueError(f"El Plan de Estudios '{plan_nombre}' no existe en el sistema.")

        # 2. Validar formato de promedio si se ingresó
        if promedio is not None:
            try:
                promedio = float(promedio)
                if not (0 <= promedio <= 100):
                    raise ValueError("El promedio debe ser un valor entre 0 y 100.")
            except Exception:
                raise ValueError(f"Promedio '{promedio}' inválido.")

        # 3. Buscar o crear Usuario (mecanismo Upsert)
        user = Usuario.objects.filter(username=control).first()
        created = False

        if not user:
            # Crear usuario nuevo
            user = Usuario(
                username=control,
                numero_control=control,
                first_name=first_name,
                last_name=last_name,
                apellido_materno=apellido_materno,
                email=email,
                rol=Rol.ALUMNO,
                carrera=carrera,
                telefono=telefono,
                genero=genero,
                generacion=generacion,
                correo_institucional=email,
                debe_cambiar_password=True  # Forzar cambio de contraseña en su primer login
            )
            # Contraseña por defecto siguiendo el patrón corporativo seguro: Ita.[Control]!
            default_pwd = f"Ita.{control}!"
            user.set_password(default_pwd)
            user.save()
            created = True

            # Registrar para enviar correo post-commit
            self.newly_created_users.append({
                'username': control,
                'email': email,
                'password': default_pwd,
                'full_name': f"{first_name} {last_name} {apellido_materno}".strip().upper()
            })
        else:
            # Actualizar campos existentes
            cambio = False
            if (user.first_name != first_name or user.last_name != last_name or
                user.apellido_materno != apellido_materno or user.email != email or
                user.carrera != carrera or user.telefono != telefono or
                user.genero != genero or user.generacion != generacion or
                user.numero_control != control):
                
                user.first_name = first_name
                user.last_name = last_name
                user.apellido_materno = apellido_materno
                user.email = email
                user.carrera = carrera
                user.telefono = telefono
                user.genero = genero
                user.generacion = generacion
                user.numero_control = control
                user.correo_institucional = email
                user.save()
                cambio = True

        # 4. Sincronizar PerfilAlumno
        perfil, p_created = PerfilAlumno.objects.get_or_create(
            usuario=user,
            defaults={
                'numero_control': control,
                'carrera': carrera,
                'plan_estudios': plan,
                'semestre_egreso': semestre_egreso,
                'promedio': promedio
            }
        )
        if not p_created:
            p_cambio = False
            if (perfil.numero_control != control or perfil.carrera != carrera or
                perfil.plan_estudios != plan or perfil.semestre_egreso != semestre_egreso or
                perfil.promedio != promedio):
                
                perfil.numero_control = control
                perfil.carrera = carrera
                perfil.plan_estudios = plan
                perfil.semestre_egreso = semestre_egreso
                perfil.promedio = promedio
                perfil.save()
                p_cambio = True
            
            if not created and (cambio or p_cambio):
                stat['actualizados'] += 1
        else:
            if not created:
                stat['actualizados'] += 1
            else:
                stat['creados'] += 1


class ValidationError(Exception):
    """Excepción controlada para forzar rollback de BD al detectar fallos lógicos."""
    pass
