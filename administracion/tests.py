import io
import openpyxl
from django.test import TestCase
from django.urls import reverse
from django.contrib.auth import get_user_model
from django.contrib.messages import get_messages

from administracion.models import Departamento, Carrera, Profesor, Rol, Genero
from expediente.models import PlanEstudios, Modalidad
from alumnos.models import PerfilAlumno

Usuario = get_user_model()


class ExcelImportExportTestCase(TestCase):
    """
    Suite de pruebas para validar la descarga de plantillas,
    la importación atómica (Upsert), el control de duplicados
    y la integridad transaccional (Rollback) con base de datos limpia.
    """

    def setUp(self):
        # Crear superusuario administrador para autenticación
        self.admin = Usuario.objects.create_superuser(
            username='admin_test',
            email='admin@apizaco.tecnm.mx',
            password='AdminPassword123!',
            rol=Rol.ADMINISTRADOR
        )
        self.client.force_login(self.admin)

    def create_in_memory_excel(self, sheets_data):
        """
        Crea un libro de openpyxl en memoria y lo retorna como un archivo para simulación de subida.
        """
        wb = openpyxl.Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        for title, headers, rows in sheets_data:
            ws = wb.create_sheet(title=title)
            ws.append(headers)
            for row in rows:
                ws.append(row)

        file_io = io.BytesIO()
        wb.save(file_io)
        file_io.seek(0)
        file_io.name = 'test_template.xlsx'
        return file_io

    def test_descargar_plantilla_vacia_ok(self):
        """
        Verifica que se generen y descarguen las plantillas vacías correctas (Todo e individuales).
        """
        tipos = ['todo', 'departamentos', 'carreras', 'planes', 'modalidades', 'profesores', 'alumnos']
        for t in tipos:
            response = self.client.get(reverse('administracion:descargar_plantilla'), {'tipo': t, 'con_datos': 'false'})
            self.assertEqual(response.status_code, 200)
            self.assertEqual(
                response['Content-Type'],
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            # Verificar estructura básica leyendo el bytes retornado
            wb = openpyxl.load_workbook(io.BytesIO(response.content))
            if t == 'todo':
                self.assertIn('Departamentos', wb.sheetnames)
                self.assertIn('Carreras', wb.sheetnames)
                self.assertIn('Alumnos', wb.sheetnames)
            elif t == 'carreras':
                self.assertIn('Carreras', wb.sheetnames)
                self.assertEqual(len(wb.sheetnames), 1)

    def test_importar_masivo_exitoso_upsert(self):
        """
        Valida que se creen registros correctamente en la base de datos limpia
        y que posteriores cargas del mismo archivo no dupliquen, sino que actualicen.
        """
        # 1. Definir los datos de importación con llaves consistentes
        data_excel = [
            (
                'Departamentos',
                ['Clave', 'Nombre del Departamento', 'Rol Responsable'],
                [
                    ['CB', 'Ciencias Basicas', 'ACADEMICO'],
                    ['CH', 'Ciencias Humanas', 'JEFE_PROYECTO']
                ]
            ),
            (
                'Carreras',
                ['Clave', 'Nombre de la Carrera', 'Activa (Si/No)', 'Clave Departamento'],
                [
                    ['ISC', 'Ingenieria en Sistemas Computacionales', 'Si', 'CB'],
                    ['IGE', 'Ingenieria en Gestion Empresarial', 'Si', 'CH']
                ]
            ),
            (
                'Planes de Estudio',
                ['Nombre (Clave)', 'Descripción', 'Activo (Si/No)'],
                [
                    ['ISIC-2010-224', 'Plan 2010 de ISC', 'Si']
                ]
            ),
            (
                'Modalidades',
                ['Plan de Estudios Nombre', 'Clave de Modalidad', 'Nombre de Modalidad', 'Descripción', 'Activa (Si/No)'],
                [
                    ['ISIC-2010-224', 'TESIS', 'Tesis Profesional', 'Examen oral', 'Si']
                ]
            ),
            (
                'Profesores',
                ['Cédula Profesional', 'Nombre(s)', 'Apellido Paterno', 'Apellido Materno', 'Título Académico', 'Email', 'Clave Departamento', 'Activo (Si/No)'],
                [
                    ['14098703', 'JUAN', 'RAMOS', 'RAMOS', 'Dr. en S.C.', 'jramos@test.com', 'CB', 'Si']
                ]
            ),
            (
                'Alumnos',
                ['Número de Control', 'Nombre(s)', 'Apellido Paterno', 'Apellido Materno', 'Email', 'Clave Carrera', 'Plan de Estudios Nombre', 'Semestre de Egreso', 'Promedio General', 'Teléfono', 'Género (M/F/O)', 'Generación (Año)'],
                [
                    ['20141720', 'DANIELA', 'SUAREZ', 'LOPEZ', 'daniela@test.com', 'ISC', 'ISIC-2010-224', 'Ago-Dic 2024', '95.50', '2411122334', 'F', 2020]
                ]
            )
        ]

        excel_file = self.create_in_memory_excel(data_excel)

        # 2. Realizar POST de subida (Modo Completo / Todo)
        response = self.client.post(
            reverse('administracion:subir_excel'),
            {'tipo': 'todo', 'archivo_excel': excel_file}
        )
        self.assertRedirects(response, reverse('administracion:importar_exportar') + '?tab=todo')

        # 3. Validar creaciones exitosas en BD
        self.assertEqual(Departamento.objects.count(), 2)
        self.assertEqual(Carrera.objects.count(), 2)
        self.assertEqual(PlanEstudios.objects.count(), 1)
        self.assertEqual(Modalidad.objects.count(), 1)
        self.assertEqual(Profesor.objects.count(), 1)
        self.assertEqual(PerfilAlumno.objects.count(), 1)

        # Validar contraseña por defecto y cambio obligatorio para Alumno
        alumno = Usuario.objects.get(username='20141720')
        self.assertEqual(alumno.rol, Rol.ALUMNO)
        self.assertTrue(alumno.debe_cambiar_password)
        self.assertTrue(alumno.check_password('Ita.20141720!'))

        # 4. Modificar datos de prueba en la misma plantilla para testear el Upsert (Modificar sin duplicar)
        data_excel_update = [
            (
                'Departamentos',
                ['Clave', 'Nombre del Departamento', 'Rol Responsable'],
                [
                    ['CB', 'Ciencias Basicas Editado', 'ACADEMICO']  # Editado
                ]
            ),
            (
                'Alumnos',
                ['Número de Control', 'Nombre(s)', 'Apellido Paterno', 'Apellido Materno', 'Email', 'Clave Carrera', 'Plan de Estudios Nombre', 'Semestre de Egreso', 'Promedio General', 'Teléfono', 'Género (M/F/O)', 'Generación (Año)'],
                [
                    ['20141720', 'DANIELA', 'SUAREZ', 'LOPEZ', 'daniela.nuevo@test.com', 'ISC', 'ISIC-2010-224', 'Ago-Dic 2024', '98.00', '2411122334', 'F', 2020]  # Email y Promedio Editados
                ]
            )
        ]

        excel_file_update = self.create_in_memory_excel(data_excel_update)

        # Subir hoja específica de Alumnos
        response_update = self.client.post(
            reverse('administracion:subir_excel'),
            {'tipo': 'alumnos', 'archivo_excel': excel_file_update}
        )
        self.assertRedirects(response_update, reverse('administracion:importar_exportar') + '?tab=alumnos')

        # Comprobar que no se duplicaron registros
        self.assertEqual(Usuario.objects.filter(rol=Rol.ALUMNO).count(), 1)
        # Comprobar que se actualizaron los campos
        perfil = PerfilAlumno.objects.get(numero_control='20141720')
        self.assertEqual(perfil.promedio, 98.00)
        self.assertEqual(perfil.correo_institucional, 'daniela.nuevo@test.com')

    def test_rollback_transaccional_por_error_referencial(self):
        """
        Valida que si se ingresa un dato erróneo (llave foránea que no existe),
        se genere un reporte y se cancele toda la carga (rollback) manteniendo limpia la base de datos.
        """
        # Definir datos donde 'Carreras' tiene un departamento 'INVALIDO' que no existe en 'Departamentos'
        data_erronea = [
            (
                'Departamentos',
                ['Clave', 'Nombre del Departamento', 'Rol Responsable'],
                [
                    ['CB', 'Ciencias Basicas', 'ACADEMICO']
                ]
            ),
            (
                'Carreras',
                ['Clave', 'Nombre de la Carrera', 'Activa (Si/No)', 'Clave Departamento'],
                [
                    ['ISC', 'Sistemas', 'Si', 'INVALIDO']  # Error: INVALIDO no fue cargado en la hoja anterior
                ]
            )
        ]

        excel_file = self.create_in_memory_excel(data_erronea)

        # Realizar POST
        response = self.client.post(
            reverse('administracion:subir_excel'),
            {'tipo': 'todo', 'archivo_excel': excel_file}
        )

        # Comprobar que no redirecciona sino que renderiza la misma página con código 200 y reporta el error
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Se canceló la importación')
        self.assertContains(response, "El Departamento con clave &#x27;INVALIDO&#x27; no está registrado")

        # Comprobar que se ejecutó el rollback de forma impecable (Departamentos sigue vacío a pesar de estar bien escrito)
        self.assertEqual(Departamento.objects.count(), 0)
        self.assertEqual(Carrera.objects.count(), 0)
